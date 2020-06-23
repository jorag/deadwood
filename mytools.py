#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 10 16:17:30 2018

@author: jorag
@author of length: MadsAdrian, updated by jorag 
@autor of ask_multiple_choice_question: SO username Kevin, updated by jorag 
"""

import numpy as np
from tkinter import Tk, Label, Button, Radiobutton, IntVar
from scipy.ndimage.filters import uniform_filter
from scipy.ndimage.measurements import variance


def ask_multiple_choice_question(prompt, options, title=None, default_v=0):
    """Multiple choice question.
    
    From: https://stackoverflow.com/questions/42581016/
    """
    root = Tk()
    if title is not None:
        root.title(title)
    if prompt:
        Label(root, text=prompt).pack()
    v = IntVar(None, default_v)
    for i, option in enumerate(options):
        Radiobutton(root, text=option, variable=v, value=i).pack(anchor="w")
    Button(text="Submit", command=root.destroy).pack()
    root.mainloop()

    return options[v.get()], v.get() 


def length(x):
    """Returns length of input.
    
    Mimics MATLAB's length function.
    """
    if isinstance(x, (int, float, complex, np.int64)):
        return 1
    elif isinstance(x, np.ndarray):
        return max(x.shape)
    try:
        return len(x)
    except TypeError as e:
        print('In length, add handling for type ' + str(type(x)))
        raise e
        
        
def numel(x):
    """Returns the number of elements in the input.
    
    Mimics MATLAB's numel function.
    """
    if isinstance(x, str):
        return 1
    elif isinstance(x, tuple):
        return 1
    else:
        return length(x)
    
    
def make_list(x):
    """Make the input into a list.
    
    Currently only for inputs with one element.
    """
    if numel(x) == 1:
        if isinstance(x, (int, float, complex, str, np.int64)):
            return [x]
        elif isinstance(x, np.ndarray):
            raise NotImplementedError('Implement support for numpy arrays in mytools.makelist(x)!')
        elif isinstance(x, tuple):
            # TODO: Should each tuple element be a list element?
            return [x]
        elif isinstance(x, list):
            return x
        else:
            raise NotImplementedError('Implement support for data type in mytools.makelist(x)!')
    else:
        raise NotImplementedError('Implement support for data with numel(x) > 1 in mytools.makelist(x)!')


def get_label_weights(labels):
    """Create a vector with weights for each label according to number of 
    occurances of each label. 
    
    Moved from dataclass.py
    """
    # Find the unique labels and counts
    unique_labels, label_counts = np.unique(labels, return_counts=True)
    # Number of points
    n_points = np.sum(label_counts)
    # Set the default weights
    weights = np.ones((length(labels),)) / n_points
    # Get relative weights of each class
    #rel_weights = label_counts/n_points
    for i_label in unique_labels:
        weights[np.where(labels==i_label)] = label_counts[i_label]/n_points 
                
    # Normalize weights and return
    weights = weights/np.sum(weights)
    return weights

        
def mynormal(x, params):
    """Univariate normal probability density function.
    
    Input:
        x - point(s)
        params - contains mu and sigma (in that order)
    Output:
        pdf(x)
    """
    mu = params[0]
    sigma = params[1]
    
    pdf = 1/(np.sqrt(2*np.pi) * sigma) * np.exp(-(x - mu)**2/(2*sigma**2))
    return pdf


def logit(txt_in, log_type='print'):
    """Function for logging text to file or printing to console.
    
    Enables a unified command for printing text during development/debugging
    and logging the text to a file during normal (mature) operations.
    """
    if log_type == 'print':
        print(txt_in)
    elif log_type == 'default':
        # TODO: Implement this (read default state)
        print(txt_in)
        
    return


def imtensor2array(input_im):
    """Convert 3D image array on the form channel, x, y to 2D array with 
    channels as columns.
    
    Useful for classification or normalization.
    """
    # Get number of channels etc.
    n_channels = input_im.shape[0]
    n_rows = input_im.shape[1]
    n_cols = input_im.shape[2]
    # Reshape array to n_cols*n_rows rows with the channels as columns 
    input_im = np.transpose(input_im, (1, 2, 0)) # Change order to rows, cols, channels
    output_im = np.reshape(input_im, (n_rows*n_cols, n_channels))
        
    return output_im, n_rows, n_cols 


def norm01(input_array, norm_type='global', min_cap=None, max_cap=None, min_cap_value=np.NaN, max_cap_value=np.NaN, log_type=None):
    """Normalize data.
    
    Parameters:
    norm_type:
        'none' - return input array (might print min, max and mean)
        'local' - min and max of each variable (column) is 0 and 1
        'global' - min of array is 0, max is 1
    min_cap: Truncate values below this value to min_cap_value before normalizing
    max_cap: Truncate values above this value to max_cap_value before normalizing
    """
    
    # Ensure that original input is not modified
    output_array = np.array(input_array, copy=True)
    
    # Replace values outside envolope/cap with NaNs (or specifie value)
    if min_cap is not None:
        output_array[output_array   < min_cap] = min_cap_value
                   
    if max_cap is not None:
        output_array[output_array  > max_cap] = max_cap_value
    
    # Normalize data for selected normalization option
    if norm_type.lower() in ['global', 'all', 'set']:
        # Normalize to 0-1 (globally)
        output_array = input_array - np.nanmin(input_array)
        output_array = output_array/np.nanmax(output_array)
    elif norm_type.lower() in ['local', 'separate']:
        # Normalize to 0-1 for each column
        output_array = input_array - np.nanmin(input_array, axis=0)
        output_array = output_array/np.nanmax(output_array, axis=0)
    elif norm_type.lower() in ['band', 'channel']:
        # Normalize to 0-1 for each channel (assumed to be last index)
        # Get shape of input
        input_shape = input_array.shape
        output_array = np.zeros(input_shape)
        # Normalize each channel
        for i_channel in range(0, input_shape[2]):
            output_array[:,:,i_channel] = input_array[:,:,i_channel] - np.nanmin(input_array[:,:,i_channel])
            output_array[:,:,i_channel] = output_array[:,:,i_channel]/np.nanmax(output_array[:,:,i_channel])
    
    # Log / print results
    if log_type is not None:
        logit('\n INPUT:', log_type)
        logit(np.nanmean(input_array), log_type)
        logit(np.nanmax(input_array, axis=0), log_type)
        logit(np.nanmean(input_array, axis=0), log_type)
        logit(np.nanmin(input_array, axis=0), log_type)
        
        logit('\n OUTPUT:', log_type)
        logit(np.nanmean(output_array), log_type)
        logit(np.nanmax(output_array, axis=0), log_type)
        logit(np.nanmean(output_array, axis=0), log_type)
        logit(np.nanmin(output_array, axis=0), log_type)
        
    return output_array


def mycolourvec(shuffle=False):
    """Get a standardized colour vector (list) for plotting.
    
    Shorthand to avoid defining it for every plot that needs different colours 
    to distinguish between classes etc.
    """
    # Define vector with colours
    colour_vec = ['r', 'g', 'b', 'm', 'c', 'y', 'k']
    
    # Randomize order?
    if shuffle:
        np.random.shuffle(colour_vec)
        
    return colour_vec


def dB(x, ref=1, input_type='power'):
    """Return result, x, in decibels (dB) relative to reference, ref."""    
    if input_type.lower() in ['power', 'pwr']:
        a = 10
    elif input_type.lower() in ['aplitude', 'amp']:
        a = 20
    return a*(np.log10(x) - np.log10(ref))


def identity(x, **kwargs):
    """Identity function that ignores extra arguments."""    
    return x 


def pauli_rgb(x):
    """Create the Pauli RGB image from input.
    
    If the number of channels in is 3 or 4, the input is assumed to be complex.
    If the number of channels in is 6 or 8, the input is assumed to be I and Q.
    (Inphase and Quadrature components in sequential band order.)
    Either way, the values are assumed to represent amplitude (not intensity).
    """
    shape_in = x.shape
    # Initialize output
    rgb_out = np.zeros((shape_in[0], shape_in[1], 3))
    # Number of bands determines form of expression
    if shape_in[2] == 4:
        # Form is complex arrays: HH, HV, VH, VV
        rgb_out[:,:,1] = 0.5*np.abs(x[:,:,1] + x[:,:,2])**2 # G
        rgb_out[:,:,0] = 0.5*np.abs(x[:,:,0] - x[:,:,3])**2 # R
        rgb_out[:,:,2] = 0.5*np.abs(x[:,:,0] + x[:,:,3])**2 # B
    elif shape_in[2] == 3:
        # Form is complex arrays: HH, HV, VV  (reciprocity assumed) 
        rgb_out[:,:,1] = 0.5*np.abs(x[:,:,1])**2 # G
        rgb_out[:,:,0] = 0.5*np.abs(x[:,:,0] - x[:,:,2])**2 # R
        rgb_out[:,:,2] = 0.5*np.abs(x[:,:,0] + x[:,:,2])**2 # B
    elif shape_in[2] == 8:
        # Form is real arrays: i_HH, q_HH, i_HV, q_HV, i_VH, q_VH, i_VV, q_VV
        rgb_out[:,:,1] = 0.5*((x[:,:,2]+x[:,:,4])**2 + (x[:,:,3]+x[:,:,5])**2) # G
        rgb_out[:,:,0] = 0.5*((x[:,:,0]-x[:,:,6])**2 + (x[:,:,1]-x[:,:,7])**2) # R
        rgb_out[:,:,2] = 0.5*((x[:,:,0]+x[:,:,6])**2 + (x[:,:,1]+x[:,:,7])**2) # B
    elif shape_in[2] == 6:
        # Form is real arrays: i_HH, q_HH, i_HV, q_HV, i_VV, q_VV (reciprocity assumed) 
        rgb_out[:,:,1] = 0.5*((x[:,:,2])**2 + (x[:,:,3])**2) # G
        rgb_out[:,:,0] = 0.5*((x[:,:,0]-x[:,:,4])**2 + (x[:,:,1]-x[:,:,5])**2) # R
        rgb_out[:,:,2] = 0.5*((x[:,:,0]+x[:,:,4])**2 + (x[:,:,1]+x[:,:,5])**2) # B
               
    return rgb_out 


def lexi_rgb(x):
    """Create the lexicographical RGB image from input.
    
    If the number of channels in is 3 or 4, the input is assumed to be complex.
    If the number of channels in is 6 or 8, the input is assumed to be I and Q.
    (Inphase and Quadrature components in sequential band order.)
    Either way, the values are assumed to represent amplitude (not intensity).
    """
    shape_in = x.shape
    # Initialize output
    rgb_out = np.zeros((shape_in[0], shape_in[1], 3))
    # Number of bands determines form of expression
    if shape_in[2] == 4:
        # Form is complex arrays: HH, HV, VH, VV
        rgb_out[:,:,0] = np.abs(x[:,:,0]) # R
        rgb_out[:,:,1] = 0.5*(np.abs(x[:,:,1]) + np.abs(x[:,:,2])) # G
        rgb_out[:,:,2] = np.abs(x[:,:,3]) # B
    elif shape_in[2] == 3:
        # Form is complex arrays: HH, HV, VV  (reciprocity assumed) 
        rgb_out[:,:,0] = np.abs(x[:,:,0])# R
        rgb_out[:,:,1] = np.abs(x[:,:,1]) # G
        rgb_out[:,:,2] = np.abs(x[:,:,2]) # B
    elif shape_in[2] == 8:
        # Form is real arrays: i_HH, q_HH, i_HV, q_HV, i_VH, q_VH, i_VV, q_VV
        rgb_out[:,:,0] = np.sqrt((x[:,:,0])**2 + (x[:,:,1])**2) # R
        rgb_out[:,:,1] = np.sqrt( (0.5*(x[:,:,2]+x[:,:,4]))**2 + (0.5*(x[:,:,3]+x[:,:,5]))**2) # G
        rgb_out[:,:,2] = np.sqrt((x[:,:,6])**2 + (x[:,:,7])**2) # B
    elif shape_in[2] == 6:
        # Form is real arrays: i_HH, q_HH, i_HV, q_HV, i_VV, q_VV (reciprocity assumed) 
        rgb_out[:,:,0] = np.sqrt((x[:,:,0])**2 + (x[:,:,1])**2) # R
        rgb_out[:,:,1] = np.sqrt((x[:,:,1])**2 + (x[:,:,2])**2) # G
        rgb_out[:,:,2] = np.sqrt((x[:,:,4])**2 + (x[:,:,5])**2) # B
               
    return rgb_out 


def iq2complex(x, reciprocity=False):
    """Merge I and Q bands to complex valued array.
    
    Create an array with complex values from separate, real-vauled Inphase and 
    Quadrature components.
    """
    shape_in = x.shape
    # Number of bands determines form of expression
    if reciprocity and shape_in[2] == 8:
        # Initialize output
        array_out = np.zeros((shape_in[0], shape_in[1], 3), dtype=complex)
        # Input is real arrays: i_HH, q_HH, i_HV, q_HV, i_VH, q_VH, i_VV, q_VV
        array_out[:,:,0] = x[:,:,0] + 1j * x[:,:,1] # HH
        array_out[:,:,1] = (x[:,:,2] + 1j*x[:,:,3] + x[:,:,4] + 1j*x[:,:,5])/2 # HV (=VH)
        array_out[:,:,2] = x[:,:,6] + 1j * x[:,:,7] # VV
    elif not reciprocity and shape_in[2] == 8:
        # Initialize output
        array_out = np.zeros((shape_in[0], shape_in[1], 4), dtype=complex)
        # Input is real arrays: i_HH, q_HH, i_HV, q_HV, i_VH, q_VH, i_VV, q_VV
        array_out[:,:,0] = x[:,:,0] + 1j * x[:,:,1] # HH
        array_out[:,:,1] = x[:,:,2] + 1j * x[:,:,3] # HV
        array_out[:,:,2] = x[:,:,4] + 1j * x[:,:,5] # VH
        array_out[:,:,3] = x[:,:,6] + 1j * x[:,:,7] # VV
    elif shape_in[2] == 6:
        # Initialize output
        array_out = np.zeros((shape_in[0], shape_in[1], 3), dtype=complex)
        # Input is real arrays: i_HH, q_HH, i_HV, q_HV, i_VV, q_VV (reciprocity assumed) 
        array_out[:,:,0] = x[:,:,0] + 1j * x[:,:,1] # HH
        array_out[:,:,1] = x[:,:,2] + 1j * x[:,:,3] # HV (=VH)
        array_out[:,:,2] = x[:,:,4] + 1j * x[:,:,5] # VH
               
    return array_out 


def polar2complex(amp_in, ang_in):
    """Create complex valued array from amplitude and angle"""

    array_out = np.zeros(amp_in.shape, dtype=complex)
    array_out = amp_in * np.exp(1j * ang_in)
               
    return array_out 


def get_sar_features(input, x_list=None, y_list=None, feature_type='not set', input_type='vec'):
    """ Get features from SAR data tensor.
    
    e.g. extract certain parameters from C3 covariance matrix.
    Works for both a full image and a list of pixels (used for training or 
    testing).
    """
    # Ensure that input is not modified
    filtered = np.copy(input)
    
    # List of data vectors
    if input_type.lower() in ['vec', 'vector', 'list']:
        if x_list is not None and y_list is not None:
            # Extract pixels
            filtered = filtered[x_list,y_list,:]
        
        # Check if current processing type indicates that a subset of features 
        # should be extracted, or if array should be returned as-is
        if feature_type.lower() in ['iq2c3']:
            temp = filtered[:, [0,2,4,8]]
            filtered = np.zeros((filtered.shape[0],5))
            filtered[:,0] = np.real(temp[:,0]) # C11
            filtered[:,1] = np.real(temp[:,2]) # C22
            filtered[:,2] = np.real(temp[:,3]) # C33
            filtered[:,3] = np.abs(temp[:,1]) # C13 abs
            filtered[:,4] = np.angle(temp[:,1]) # C13 angle
        elif feature_type.lower() in ['c3_snap_5feat', 'c3snap_filtered', 'c3snap5feat']:
            temp = filtered[:, [0,3,4,5,8]]
            filtered = np.zeros((filtered.shape[0],5))
            filtered[:,0] = temp[:,0] # C11
            filtered[:,1] = temp[:,3] # C22
            filtered[:,2] = temp[:,4] # C33
            # Make complex values and use abs and angle formulas as for iq2c3 
            # (ensure same angle calculation)
            filtered[:,3] = np.abs(temp[:,1]+ 1j* temp[:,2]) 
            filtered[:,4] = np.angle(temp[:,1]+ 1j* temp[:,2])
        elif feature_type.lower() in ['c3_pgnlm2intensities','c3pgnlm5feat2intensities']:
            filtered = filtered[:, [0,1,2]]
        elif feature_type.lower() in ['c3_pgnlm','c3pgnlm5feat', 'c3pgnlm5feat25feat']:
            pass
        elif feature_type.lower() in ['same', 'all']:
            pass 
        else:
            print('WARNING! Feature type '+feature_type+ ' not defined in get_sar_features!!')
    
    # Image array of data vectors
    elif input_type.lower() in ['img', 'image', 'full']:
        # Check if current processing type indicates that a subset of features 
        # should be extracted, or if array should be returned as-is
        if feature_type.lower() in ['iq2c3']:
            temp = filtered[:,:, [0,2,4,8]]
            filtered = np.zeros((filtered.shape[0], filtered.shape[1], 5))
            filtered[:,:,0] = np.real(temp[:,:,0]) # C11
            filtered[:,:,1] = np.real(temp[:,:,2]) # C22
            filtered[:,:,2] = np.real(temp[:,:,3]) # C33
            filtered[:,:,3] = np.abs(temp[:,:,1]) # C13 abs
            filtered[:,:,4] = np.angle(temp[:,:,1]) # C13 angle
        elif feature_type.lower() in ['c3_snap_5feat','c3snap_filtered', 'c3snap5feat']:
            temp = filtered[:,:, [0,3,4,5,8]]
            filtered = np.zeros((filtered.shape[0], filtered.shape[1], 5))
            filtered[:,:,0] = temp[:,:,0] # C11
            filtered[:,:,1] = temp[:,:,3] # C22
            filtered[:,:,2] = temp[:,:,4] # C33
            # Make complex values and use abs and angle formulas as for iq2c3 
            # (ensure same angle calculation)
            filtered[:,:,3] = np.abs(temp[:,:,1]+ 1j* temp[:,:,2]) 
            filtered[:,:,4] = np.angle(temp[:,:,1]+ 1j* temp[:,:,2])
        elif feature_type.lower() in ['c3_snap_intensities','c3snap2intensities']:
            temp = filtered[:,:, [0,3,4,5,8]]
            filtered = np.zeros((filtered.shape[0], filtered.shape[1], 3))
            filtered[:,:,0] = temp[:,:,0] # C11
            filtered[:,:,1] = temp[:,:,3] # C22
            filtered[:,:,2] = temp[:,:,4] # C33
        elif feature_type.lower() in ['c3_pgnlm2intensities','c3pgnlm5feat2intensities']:
            filtered = filtered[:,:, [0,1,2]]
        elif feature_type.lower() in ['c3_pgnlm','c3pgnlm5feat', 'c3pgnlm5feat25feat']:
            pass 
        elif feature_type.lower() in ['abs']:
            filtered = np.abs(filtered, dtype='double') 
        elif feature_type.lower() in ['same', 'all']:
            pass 
        else:
            print('WARNING! Feature type '+feature_type+ ' not defined in get_sar_features!!')
    
    return filtered


def lee_filter(img, size):
    """ Lee filter for SAR despeckling.
    
    From Alex I.'s answer here:
    https://stackoverflow.com/questions/39785970/speckle-lee-filter-in-python
    """
    img_mean = uniform_filter(img, (size, size))
    img_sqr_mean = uniform_filter(img**2, (size, size))
    img_variance = img_sqr_mean - img_mean**2

    overall_variance = variance(img)

    img_weights = img_variance / (img_variance + overall_variance)
    img_output = img_mean + img_weights * (img - img_mean)
    return img_outpu


def boxcar_filter(img, size):
    """ Boxcar filter."""
    img_mean = uniform_filter(img, (size, size, 1), mode='reflect')
    return img_mean


def rsquare(y, y_hat):
    """Calculate the goodness of fit, R²
    
    Assume response / labels / y is given as columns.
    """
    # Average/sum over columns
    if length(y.shape) == 1:
        axis = None
    else:
        axis = 0
    # Relative square error: SS_residual / SS_total (SS= Sum of Squares) 
    SS_residual = np.sum((y - y_hat)**2, axis=axis)
    SS_total = np.sum((y-np.mean(y, axis=axis))**2, axis=axis)
    Erse = SS_residual/SS_total
    # Coefficient Of Determination, R²
    Rsquare = 1 - Erse
    return Rsquare


class empty_object:
    """Empty object for storing and organizing parameters etc."""
    def __init__(self, **kwargs):
        # Get keyword arguments
        for key, value in kwargs.items():
            setattr(self, key, value)
            
    def print(self):
        # Print
        # 20191207 - TODO - implement this!
        for key in self.all_keys:
            print(key, ' : ', getattr(self, key))
            
            
class params_object:
    """Object for storing and organizing parameters etc."""
    def __init__(self, **kwargs):
        #self.all_keys = []
        #self.primary_keys = []
        # Get keyword arguments
        for key, value in kwargs.items():
            setattr(self, key, value)
            #self.all_keys.append(key)
            #self.primary_keys.append(key)
            
    def add(self, key, value, primary_keys = True):
        setattr(self, key, value)
#        self.all_keys.append(key)
#        if primary_keys:
#            self.primary_keys.append(key)
    
    def print_params(self, primary_keys = False):
        # Print
        pass
#        if primary_keys:
#            for key in self.primary_keys:
#                print(key, ' : ', getattr(self, key))
#        else:
#            for key in self.all_keys:
#                print(key, ' : ', getattr(self, key))
            

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # From MaxU's answer here: https://stackoverflow.com/questions/20219254/
    import pandas as pd
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
